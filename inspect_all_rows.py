import openpyxl

wb = openpyxl.load_workbook(r"D:\产品图片\网图\iColorPacks_SEO_AEO_4%E5%91%A8%E6%89%A7%E8%A1%8C%E8%AE%A1%E5%88%92.xlsx")
sheet = wb["4周执行计划"]
print("Total rows:", sheet.max_row)
for i, r in enumerate(sheet.iter_rows(values_only=True), 1):
    print(f"Row {i:2d}: {r}")
