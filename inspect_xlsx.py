import openpyxl

wb = openpyxl.load_workbook(r"D:\产品图片\网图\iColorPacks_SEO_AEO_4%E5%91%A8%E6%89%A7%E8%A1%8C%E8%AE%A1%E5%88%92.xlsx")
print("Sheets:", wb.sheetnames)
for name in wb.sheetnames:
    sheet = wb[name]
    print(f"\n--- Sheet: {name} ---")
    # Print first 20 rows
    for r in list(sheet.iter_rows(values_only=True))[:20]:
        if any(r): # skip empty rows
            print(r)
