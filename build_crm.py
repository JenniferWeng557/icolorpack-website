import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.worksheet.datavalidation import DataValidation

# Append XLSX skill scripts path for styles and builders
sys.path.append(r"C:\Users\uu\.accio\accounts\1747554937\agents\DID-F456DA-2B0D4C\agent-core\skills\xlsx\scripts")
import style_kit
import builders

# Setup paths and data sources
old_file_path = r"C:\Users\uu\Desktop\客户跟进与管理表格.xls"
new_file_path = r"C:\Users\uu\Desktop\客户跟进与管理系统_iColorPack.xlsx"

print("Reading old customer data...")
xl = pd.ExcelFile(old_file_path)

# Extract follow-up data from sheet '4' and sheet '7'
logs_data = []

# Customer 4 follow-up logs
df_4 = xl.parse('4', header=None)
# Rows 9 to 12 contain the follow-up history (index 9 to 12)
for idx in range(9, len(df_4)):
    date_val = df_4.iloc[idx, 0]
    feedback_val = df_4.iloc[idx, 1]
    remark_val = df_4.iloc[idx, 3] if df_4.shape[1] > 3 else ""
    if pd.notnull(date_val) and pd.notnull(feedback_val):
        logs_data.append({
            "customer_id": "KH0004",
            "date": date_val,
            "feedback": feedback_val,
            "remark": remark_val if pd.notnull(remark_val) else "",
            "operator": "Jennifer"
        })

# Customer 7 follow-up logs
df_7 = xl.parse('7', header=None)
# Rows 9 to 13 contain the follow-up history
for idx in range(9, len(df_7)):
    date_val = df_7.iloc[idx, 0]
    feedback_val = df_7.iloc[idx, 1]
    remark_val = df_7.iloc[idx, 3] if df_7.shape[1] > 3 else ""
    if pd.notnull(date_val) and pd.notnull(feedback_val):
        logs_data.append({
            "customer_id": "KH0007",
            "date": date_val,
            "feedback": feedback_val,
            "remark": remark_val if pd.notnull(remark_val) else "",
            "operator": "Jennifer"
        })

# Sort logs by date (ascending)
logs_data.sort(key=lambda x: x["date"])

print(f"Extracted {len(logs_data)} follow-up logs.")

# Initialize the new workbook
wb = openpyxl.Workbook()

# Set CJK font for normal style
wb._named_styles['Normal'].font = Font(name="Microsoft YaHei", size=11)

# Setup fonts and styles
FONT_NAME = "Microsoft YaHei"
font_title = Font(name=FONT_NAME, size=16, bold=True, color="C9A84C")
font_section = Font(name=FONT_NAME, size=12, bold=True, color="1C1C1C")
font_header = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
font_bold = Font(name=FONT_NAME, size=10, bold=True, color="1C1C1C")
font_normal = Font(name=FONT_NAME, size=10, color="333333")
font_kpi_num = Font(name=FONT_NAME, size=22, bold=True, color="C9A84C")
font_kpi_label = Font(name=FONT_NAME, size=9, bold=False, color="666666")

fill_header = PatternFill(start_color="1C1C1C", end_color="1C1C1C", fill_type="solid") # Charcoal Black
fill_kpi = PatternFill(start_color="F9F6F0", end_color="F9F6F0", fill_type="solid") # Light Cream/Gold
fill_accent = PatternFill(start_color="C9A84C", end_color="C9A84C", fill_type="solid") # Champagne Gold
fill_band = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid") # Clean white-grey

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

double_bottom_border = Border(
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='double', color='1C1C1C')
)

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)

# ---------------------------------------------------------------------------
# Sheet 1: Dashboard (数据看板)
# ---------------------------------------------------------------------------
ws_dash = wb.active
ws_dash.title = "数据看板"
ws_dash.views.sheetView[0].showGridLines = True

# Title
ws_dash["A2"] = "iColorPack 客户管理与跟进数据看板"
ws_dash["A2"].font = font_title
ws_dash["A3"] = "最后更新时间: 2026-06-09 (动态计算)"
ws_dash["A3"].font = Font(name=FONT_NAME, size=10, italic=True, color="666666")

# Set up KPI Cards in Row 5-7
# Card 1: 总客户数
ws_dash.merge_cells("B5:C5")
ws_dash["B5"] = "总客户数"
ws_dash["B5"].font = font_kpi_label
ws_dash["B5"].alignment = align_center
ws_dash["B5"].fill = fill_kpi

ws_dash.merge_cells("B6:C7")
ws_dash["B6"] = '=COUNTIFS(客户跟进总表!C5:C104, "<>")'
ws_dash["B6"].font = font_kpi_num
ws_dash["B6"].alignment = align_center
ws_dash["B6"].fill = fill_kpi

for r in range(5, 8):
    for c in range(2, 4):
        ws_dash.cell(row=r, column=c).border = thin_border

# Card 2: 跟进中客户
ws_dash.merge_cells("E5:F5")
ws_dash["E5"] = "跟进中客户"
ws_dash["E5"].font = font_kpi_label
ws_dash["E5"].alignment = align_center
ws_dash["E5"].fill = fill_kpi

ws_dash.merge_cells("E6:F7")
ws_dash["E6"] = '=COUNTIFS(客户跟进总表!B5:B104, "<>流失客户", 客户跟进总表!C5:C104, "<>")'
ws_dash["E6"].font = font_kpi_num
ws_dash["E6"].alignment = align_center
ws_dash["E6"].fill = fill_kpi

for r in range(5, 8):
    for c in range(5, 7):
        ws_dash.cell(row=r, column=c).border = thin_border

# Card 3: 近30天跟进客户
ws_dash.merge_cells("H5:I5")
ws_dash["H5"] = "30天内已跟进"
ws_dash["H5"].font = font_kpi_label
ws_dash["H5"].alignment = align_center
ws_dash["H5"].fill = fill_kpi

ws_dash.merge_cells("H6:I7")
ws_dash["H6"] = '=COUNTIFS(客户跟进总表!O5:O104, "<=30", 客户跟进总表!C5:C104, "<>")'
ws_dash["H6"].font = font_kpi_num
ws_dash["H6"].alignment = align_center
ws_dash["H6"].fill = fill_kpi

for r in range(5, 8):
    for c in range(8, 10):
        ws_dash.cell(row=r, column=c).border = thin_border

# Card 4: 超期/今日待跟进
ws_dash.merge_cells("K5:L5")
ws_dash["K5"] = "待跟进提醒"
ws_dash["K5"].font = font_kpi_label
ws_dash["K5"].alignment = align_center
ws_dash["K5"].fill = fill_kpi

ws_dash.merge_cells("K6:L7")
ws_dash["K6"] = '=COUNTIFS(客户跟进总表!Q5:Q104, "*待跟进*", 客户跟进总表!C5:C104, "<>") + COUNTIFS(客户跟进总表!Q5:Q104, "*超期*", 客户跟进总表!C5:C104, "<>")'
ws_dash["K6"].font = font_kpi_num
ws_dash["K6"].alignment = align_center
ws_dash["K6"].fill = fill_kpi

for r in range(5, 8):
    for c in range(11, 13):
        ws_dash.cell(row=r, column=c).border = thin_border

# Summary Table 1: 合作阶段分布 (Row 10 to 18)
ws_dash["B10"] = "合作阶段"
ws_dash["C10"] = "客户数量"
ws_dash["B10"].font = font_header
ws_dash["B10"].fill = fill_header
ws_dash["B10"].alignment = align_center
ws_dash["C10"].font = font_header
ws_dash["C10"].fill = fill_header
ws_dash["C10"].alignment = align_center

stages = ["未联系", "初步接洽", "需求沟通", "寄样确认", "报价谈判", "签约合作", "流失/归档"]
for i, stage in enumerate(stages):
    row_idx = 11 + i
    ws_dash.cell(row=row_idx, column=2, value=stage).font = font_normal
    ws_dash.cell(row=row_idx, column=2).alignment = align_left
    ws_dash.cell(row=row_idx, column=2).border = thin_border
    
    formula = f'=COUNTIFS(客户跟进总表!$L$5:$L$104, B{row_idx}, 客户跟进总表!$C$5:$C$104, "<>")'
    ws_dash.cell(row=row_idx, column=3, value=formula).font = font_normal
    ws_dash.cell(row=row_idx, column=3).alignment = align_right
    ws_dash.cell(row=row_idx, column=3).border = thin_border
    ws_dash.cell(row=row_idx, column=3).number_format = '#,##0'

ws_dash["B18"] = "总计"
ws_dash["B18"].font = font_bold
ws_dash["B18"].border = double_bottom_border
ws_dash["C18"] = "=SUM(C11:C17)"
ws_dash["C18"].font = font_bold
ws_dash["C18"].alignment = align_right
ws_dash["C18"].border = double_bottom_border
ws_dash["C18"].number_format = '#,##0'


# Summary Table 2: 客户来源分布 (Row 10 to 19)
ws_dash["E10"] = "客户来源"
ws_dash["F10"] = "客户数量"
ws_dash["E10"].font = font_header
ws_dash["E10"].fill = fill_header
ws_dash["E10"].alignment = align_center
ws_dash["F10"].font = font_header
ws_dash["F10"].fill = fill_header
ws_dash["F10"].alignment = align_center

sources = ["RFQ主动报价", "TM询盘", "自主开发", "抖音", "TIKTOK", "1688", "香港展会", "广交会"]
for i, src in enumerate(sources):
    row_idx = 11 + i
    ws_dash.cell(row=row_idx, column=5, value=src).font = font_normal
    ws_dash.cell(row=row_idx, column=5).alignment = align_left
    ws_dash.cell(row=row_idx, column=5).border = thin_border
    
    formula = f'=COUNTIFS(客户跟进总表!$J$5:$J$104, E{row_idx}, 客户跟进总表!$C$5:$C$104, "<>")'
    ws_dash.cell(row=row_idx, column=6, value=formula).font = font_normal
    ws_dash.cell(row=row_idx, column=6).alignment = align_right
    ws_dash.cell(row=row_idx, column=6).border = thin_border
    ws_dash.cell(row=row_idx, column=6).number_format = '#,##0'

ws_dash["E19"] = "总计"
ws_dash["E19"].font = font_bold
ws_dash["E19"].border = double_bottom_border
ws_dash["F19"] = "=SUM(F11:F18)"
ws_dash["F19"].font = font_bold
ws_dash["F19"].alignment = align_right
ws_dash["F19"].border = double_bottom_border
ws_dash["F19"].number_format = '#,##0'

# ---------------------------------------------------------------------------
# Sheet 2: 客户跟进总表 (Customer Follow-up Master)
# ---------------------------------------------------------------------------
ws_master = wb.create_sheet("客户跟进总表")
ws_master.views.sheetView[0].showGridLines = True

headers_master = [
    "客户ID", "客户状态", "公司名称", "国家/地区", "意向产品", "联系人", "手机号/WhatsApp", 
    "邮箱", "网址", "客户来源", "层级", "合作阶段", "首次跟进日期", "最后跟进日期", 
    "距离上次跟进(天)", "下次跟进日期", "跟进提醒", "最新跟进反馈", "最新跟进人", "备注"
]

# Style headers
for col_idx, header in enumerate(headers_master):
    cell = ws_master.cell(row=4, column=col_idx + 1, value=header)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

ws_master["A2"] = "iColorPack 客户跟进与管理总表"
ws_master["A2"].font = font_title

# Fill 100 customer rows
for i in range(1, 101):
    row_idx = 4 + i
    cust_id = f"KH{i:04d}"
    
    # Customer ID (static)
    ws_master.cell(row=row_idx, column=1, value=cust_id).font = font_bold
    ws_master.cell(row=row_idx, column=1).alignment = align_center
    ws_master.cell(row=row_idx, column=1).border = thin_border
    
    # Active records fill
    if i == 4:
        # Customer 4
        ws_master.cell(row=row_idx, column=2, value="正常跟进").font = font_normal
        ws_master.cell(row=row_idx, column=3, value="Customer 4 (温州iColor)").font = font_bold
        ws_master.cell(row=row_idx, column=4, value="中国").font = font_normal
        ws_master.cell(row=row_idx, column=10, value="自主开发").font = font_normal
        ws_master.cell(row=row_idx, column=11, value="C-良好").font = font_normal
        ws_master.cell(row=row_idx, column=12, value="初步接洽").font = font_normal
        ws_master.cell(row=row_idx, column=13, value="2023-04-14").font = font_normal
    elif i == 7:
        # Customer 7
        ws_master.cell(row=row_idx, column=2, value="正常跟进").font = font_normal
        ws_master.cell(row=row_idx, column=3, value="Customer 7 (欧美经销商)").font = font_bold
        ws_master.cell(row=row_idx, column=4, value="美国").font = font_normal
        ws_master.cell(row=row_idx, column=10, value="1688").font = font_normal
        ws_master.cell(row=row_idx, column=11, value="D-一般").font = font_normal
        ws_master.cell(row=row_idx, column=12, value="需求沟通").font = font_normal
        ws_master.cell(row=row_idx, column=13, value="2023-04-12").font = font_normal
    else:
        # Empty cells styled
        for col_idx in range(2, 14):
            ws_master.cell(row=row_idx, column=col_idx).font = font_normal
            ws_master.cell(row=row_idx, column=col_idx).alignment = align_left
            ws_master.cell(row=row_idx, column=col_idx).border = thin_border
            
    # Center style for status and dropdown columns
    for col_idx in [2, 10, 11, 12]:
        ws_master.cell(row=row_idx, column=col_idx).alignment = align_center

    # Apply styling for text fields
    for col_idx in [1, 13]:
        ws_master.cell(row=row_idx, column=col_idx).alignment = align_center

    # Column N: Last Follow-up Date (最后跟进日期)
    formula_last_date = f'=IFERROR(MAXIFS(跟进记录明细!D:D, 跟进记录明细!B:B, A{row_idx}), "")'
    ws_master.cell(row=row_idx, column=14, value=formula_last_date).font = font_normal
    ws_master.cell(row=row_idx, column=14).alignment = align_center
    ws_master.cell(row=row_idx, column=14).border = thin_border
    ws_master.cell(row=row_idx, column=14).number_format = 'yyyy-mm-dd'

    # Column O: Days Since Last Follow-up (距离上次跟进(天))
    formula_days = f'=IF(OR(N{row_idx}="", N{row_idx}=0), "", TODAY()-N{row_idx})'
    ws_master.cell(row=row_idx, column=15, value=formula_days).font = font_normal
    ws_master.cell(row=row_idx, column=15).alignment = align_right
    ws_master.cell(row=row_idx, column=15).border = thin_border
    ws_master.cell(row=row_idx, column=15).number_format = '#,##0'

    # Column P: Next Follow-up Date (下次跟进日期) - default empty but styled
    ws_master.cell(row=row_idx, column=16).font = font_normal
    ws_master.cell(row=row_idx, column=16).alignment = align_center
    ws_master.cell(row=row_idx, column=16).border = thin_border
    ws_master.cell(row=row_idx, column=16).number_format = 'yyyy-mm-dd'

    # Column Q: 跟进提醒 (Reminder Status)
    formula_remind = f'=IF(P{row_idx}="","",IF(P{row_idx}<TODAY(),"🔴 超期未跟进",IF(P{row_idx}=TODAY(),"🟡 今日需跟进","🟢 正常") ))'
    ws_master.cell(row=row_idx, column=17, value=formula_remind).font = font_normal
    ws_master.cell(row=row_idx, column=17).alignment = align_center
    ws_master.cell(row=row_idx, column=17).border = thin_border

    # Column R: 最新跟进反馈 (Latest Feedback)
    formula_feedback = f'=IFERROR(INDEX(跟进记录明细!F:F, MATCH(A{row_idx}&"_"&TEXT(N{row_idx},"yyyy-mm-dd"), 跟进记录明细!J:J, 0)), "")'
    ws_master.cell(row=row_idx, column=18, value=formula_feedback).font = font_normal
    ws_master.cell(row=row_idx, column=18).alignment = align_left
    ws_master.cell(row=row_idx, column=18).border = thin_border

    # Column S: 最新跟进人 (Latest Operator)
    formula_operator = f'=IFERROR(INDEX(跟进记录明细!H:H, MATCH(A{row_idx}&"_"&TEXT(N{row_idx},"yyyy-mm-dd"), 跟进记录明细!J:J, 0)), "")'
    ws_master.cell(row=row_idx, column=19, value=formula_operator).font = font_normal
    ws_master.cell(row=row_idx, column=19).alignment = align_center
    ws_master.cell(row=row_idx, column=19).border = thin_border

    # Column T: 备注
    ws_master.cell(row=row_idx, column=20).font = font_normal
    ws_master.cell(row=row_idx, column=20).alignment = align_left
    ws_master.cell(row=row_idx, column=20).border = thin_border

    # Zebra striping for even rows
    if i % 2 == 0:
        for col_idx in range(1, 21):
            ws_master.cell(row=row_idx, column=col_idx).fill = fill_band

ws_master.freeze_panes = "D5"

# ---------------------------------------------------------------------------
# Sheet 3: 跟进记录明细 (Follow-up Details)
# ---------------------------------------------------------------------------
ws_detail = wb.create_sheet("跟进记录明细")
ws_detail.views.sheetView[0].showGridLines = True

headers_detail = [
    "记录ID", "客户ID", "公司名称", "跟进日期", "跟进方式", "跟进反馈", "下次跟进计划", "跟进人", "备注", "辅助Key"
]

# Style headers
for col_idx, header in enumerate(headers_detail):
    cell = ws_detail.cell(row=4, column=col_idx + 1, value=header)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

ws_detail["A2"] = "iColorPack 客户跟进记录明细"
ws_detail["A2"].font = font_title

# Fill existing follow-up records
for i, log in enumerate(logs_data):
    row_idx = 5 + i
    log_id = f"JL{i+1:04d}"
    
    ws_detail.cell(row=row_idx, column=1, value=log_id).font = font_bold
    ws_detail.cell(row=row_idx, column=1).alignment = align_center
    ws_detail.cell(row=row_idx, column=1).border = thin_border
    
    ws_detail.cell(row=row_idx, column=2, value=log["customer_id"]).font = font_bold
    ws_detail.cell(row=row_idx, column=2).alignment = align_center
    ws_detail.cell(row=row_idx, column=2).border = thin_border
    
    # Auto VLOOKUP for company name
    formula_comp = f'=IFERROR(VLOOKUP(B{row_idx}, 客户跟进总表!A:C, 3, FALSE), "")'
    ws_detail.cell(row=row_idx, column=3, value=formula_comp).font = font_normal
    ws_detail.cell(row=row_idx, column=3).alignment = align_left
    ws_detail.cell(row=row_idx, column=3).border = thin_border
    
    # Date formatting
    date_str = pd.to_datetime(log["date"]).strftime("%Y-%m-%d")
    ws_detail.cell(row=row_idx, column=4, value=date_str).font = font_normal
    ws_detail.cell(row=row_idx, column=4).alignment = align_center
    ws_detail.cell(row=row_idx, column=4).border = thin_border
    ws_detail.cell(row=row_idx, column=4).number_format = 'yyyy-mm-dd'
    
    # Check method (guess based on feedback or default)
    method = "邮件跟进"
    if "微信" in log["feedback"]:
        method = "微信沟通"
    elif "加微信" in log["feedback"]:
        method = "微信沟通"
    elif "展会" in log["feedback"]:
        method = "展会面谈"
        
    ws_detail.cell(row=row_idx, column=5, value=method).font = font_normal
    ws_detail.cell(row=row_idx, column=5).alignment = align_center
    ws_detail.cell(row=row_idx, column=5).border = thin_border
    
    # Feedback
    ws_detail.cell(row=row_idx, column=6, value=log["feedback"]).font = font_normal
    ws_detail.cell(row=row_idx, column=6).alignment = align_left
    ws_detail.cell(row=row_idx, column=6).border = thin_border
    
    # Next plan
    ws_detail.cell(row=row_idx, column=7, value="").font = font_normal
    ws_detail.cell(row=row_idx, column=7).alignment = align_left
    ws_detail.cell(row=row_idx, column=7).border = thin_border
    
    # Operator
    ws_detail.cell(row=row_idx, column=8, value=log["operator"]).font = font_normal
    ws_detail.cell(row=row_idx, column=8).alignment = align_center
    ws_detail.cell(row=row_idx, column=8).border = thin_border
    
    # Remarks
    ws_detail.cell(row=row_idx, column=9, value=log["remark"]).font = font_normal
    ws_detail.cell(row=row_idx, column=9).alignment = align_left
    ws_detail.cell(row=row_idx, column=9).border = thin_border
    
    # Helper Key (辅助Key) = B5 & "_" & TEXT(D5, "yyyy-mm-dd")
    formula_key = f'=B{row_idx}&"_"&TEXT(D{row_idx},"yyyy-mm-dd")'
    ws_detail.cell(row=row_idx, column=10, value=formula_key).font = font_normal
    ws_detail.cell(row=row_idx, column=10).alignment = align_center
    ws_detail.cell(row=row_idx, column=10).border = thin_border

# Fill extra 100 empty log rows for future logging
for i in range(len(logs_data), 150):
    row_idx = 5 + i
    log_id = f"JL{i+1:04d}"
    
    ws_detail.cell(row=row_idx, column=1, value=log_id).font = font_bold
    ws_detail.cell(row=row_idx, column=1).alignment = align_center
    ws_detail.cell(row=row_idx, column=1).border = thin_border
    
    ws_detail.cell(row=row_idx, column=2).font = font_bold
    ws_detail.cell(row=row_idx, column=2).alignment = align_center
    ws_detail.cell(row=row_idx, column=2).border = thin_border
    
    formula_comp = f'=IFERROR(VLOOKUP(B{row_idx}, 客户跟进总表!A:C, 3, FALSE), "")'
    ws_detail.cell(row=row_idx, column=3, value=formula_comp).font = font_normal
    ws_detail.cell(row=row_idx, column=3).alignment = align_left
    ws_detail.cell(row=row_idx, column=3).border = thin_border
    
    for col_idx in [4, 5, 6, 7, 8, 9]:
        ws_detail.cell(row=row_idx, column=col_idx).font = font_normal
        ws_detail.cell(row=row_idx, column=col_idx).border = thin_border
        
    ws_detail.cell(row=row_idx, column=4).alignment = align_center
    ws_detail.cell(row=row_idx, column=4).number_format = 'yyyy-mm-dd'
    ws_detail.cell(row=row_idx, column=5).alignment = align_center
    ws_detail.cell(row=row_idx, column=8).alignment = align_center
    
    formula_key = f'=IF(B{row_idx}="","",B{row_idx}&"_"&TEXT(D{row_idx},"yyyy-mm-dd"))'
    ws_detail.cell(row=row_idx, column=10, value=formula_key).font = font_normal
    ws_detail.cell(row=row_idx, column=10).alignment = align_center
    ws_detail.cell(row=row_idx, column=10).border = thin_border
    
    # Zebra striping
    if i % 2 == 1:
        for col_idx in range(1, 11):
            ws_detail.cell(row=row_idx, column=col_idx).fill = fill_band

ws_detail.freeze_panes = "D5"

# ---------------------------------------------------------------------------
# Sheet 4: 参数配置 (Dropdown Settings)
# ---------------------------------------------------------------------------
ws_param = wb.create_sheet("参数配置")
ws_param.views.sheetView[0].showGridLines = True

headers_param = ["客户状态", "客户来源", "层级", "合作阶段", "跟进方式"]
for col_idx, header in enumerate(headers_param):
    cell = ws_param.cell(row=1, column=col_idx + 1, value=header)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

params_data = {
    "客户状态": ["正常跟进", "重点跟进", "暂停跟进", "已签单", "流失客户"],
    "客户来源": ["RFQ主动报价", "TM询盘", "自主开发", "抖音", "TIKTOK", "1688", "香港展会", "广交会"],
    "层级": ["A-非常优质", "B-优质", "C-良好", "D-一般", "E-较差"],
    "合作阶段": ["未联系", "初步接洽", "需求沟通", "寄样确认", "报价谈判", "签约合作", "流失/归档"],
    "跟进方式": ["邮件跟进", "WhatsApp", "电话拜访", "展会面谈", "微信沟通", "在线询盘(Alibaba)", "其他"]
}

for col_idx, header in enumerate(headers_param):
    items = params_data[header]
    for row_idx, item in enumerate(items):
        cell = ws_param.cell(row=row_idx + 2, column=col_idx + 1, value=item)
        cell.font = font_normal
        cell.alignment = align_center
        cell.border = thin_border

# ---------------------------------------------------------------------------
# Apply Dropdown Validations
# ---------------------------------------------------------------------------
print("Adding validations...")
# Validate '客户状态' (Column B of Master, dropdown from Param col A)
dv_status = DataValidation(type="list", formula1="'参数配置'!$A$2:$A$6", allow_blank=True)
ws_master.add_data_validation(dv_status)
dv_status.add("B5:B104")

# Validate '客户来源' (Column J of Master, dropdown from Param col B)
dv_source = DataValidation(type="list", formula1="'参数配置'!$B$2:$B$9", allow_blank=True)
ws_master.add_data_validation(dv_source)
dv_source.add("J5:J104")

# Validate '层级' (Column K of Master, dropdown from Param col C)
dv_level = DataValidation(type="list", formula1="'参数配置'!$C$2:$C$6", allow_blank=True)
ws_master.add_data_validation(dv_level)
dv_level.add("K5:K104")

# Validate '合作阶段' (Column L of Master, dropdown from Param col D)
dv_stage = DataValidation(type="list", formula1="'参数配置'!$D$2:$D$8", allow_blank=True)
ws_master.add_data_validation(dv_stage)
dv_stage.add("L5:L104")

# Validate '跟进方式' (Column E of Detail, dropdown from Param col E)
dv_method = DataValidation(type="list", formula1="'参数配置'!$E$2:$E$8", allow_blank=True)
ws_detail.add_data_validation(dv_method)
dv_method.add("E5:E154")

# Validate '客户ID' (Column B of Detail, dropdown from Master col A)
dv_cust_id = DataValidation(type="list", formula1="'客户跟进总表'!$A$5:$A$104", allow_blank=True)
ws_detail.add_data_validation(dv_cust_id)
dv_cust_id.add("B5:B154")

# ---------------------------------------------------------------------------
# Add Visual Charts to Dashboard
# ---------------------------------------------------------------------------
print("Building charts...")
# Chart 1: BarChart for Cooperation Stage Distribution
builders.build_chart(
    ws_dash,
    kind="clustered_bar",
    data_range="'数据看板'!B10:C17",
    title="客户合作阶段分布",
    anchor="H10",
    size=(480, 280),
    theme="executive_report",
    legend="none"
)

# Chart 2: BarChart for Customer Source Distribution
builders.build_chart(
    ws_dash,
    kind="clustered_bar",
    data_range="'数据看板'!E10:F18",
    title="客户渠道来源分布",
    anchor="N10",
    size=(480, 280),
    theme="executive_report",
    legend="none"
)

# ---------------------------------------------------------------------------
# Dynamic Column Widths
# ---------------------------------------------------------------------------
print("Adjusting widths...")
for ws in [ws_dash, ws_master, ws_detail, ws_param]:
    for col in ws.columns:
        # Default fallback
        max_len = 10
        for cell in col:
            # Skip cells with long titles, charts, or notes
            if cell.row < 4 and ws != ws_param:
                continue
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

# Save file
print("Saving new CRM spreadsheet...")
wb.save(new_file_path)
print(f"CRM system spreadsheet saved successfully at: {new_file_path}")
